import pytest
import requests
from utilities.logger import Loggen
import openpyxl

file_path = r'TestData\post_data.xlsx'

wb = openpyxl.load_workbook(file_path)

sh = wb.active

row_min = sh.min_row
col_min = sh.min_column

row_max = sh.max_row
col_max = sh.max_column

t1 = []
t2 = []

for i in range(2, row_max + 1):
    t1.append(sh.cell(i, 1).value)
    t2.append(sh.cell(i, 2).value)

l = zip(tuple(t1), tuple(t2))

k = tuple(l)

site_url = 'https://todo.pixegami.io'


@pytest.mark.parametrize("id, name", k)
def test_junk(id, name):
    print(id)
    print(name)
    res = requests.get(url=site_url)
    print(res.content)

